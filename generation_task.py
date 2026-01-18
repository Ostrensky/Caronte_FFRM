# --- FILE: app/generation_task.py ---

import os
import pandas as pd
import logging
import traceback
import time
import openpyxl 
import multiprocessing 
from queue import Empty

# Importações de lógica de geração
from main import (
    load_and_prepare_invoices, 
    perform_rules_analysis, 
    perform_description_analysis,
    load_activity_data 
)
from data_loader import create_context_for_generation
from report_generator import generate_simple_document, generate_report
from pdf_reports_generator import generate_detailed_pdfs

# Importações de configuração
from app.config import (
    get_aliquotas_path, 
    get_template_inicio_path,
    get_template_relatorio_path, 
    get_template_encerramento_dec_path,
    get_template_encerramento_ar_path,
    get_output_dir
)

def _write_to_activity_log(queue, context):
    """
    Appends a summary of the generated process to the activity log Excel file.
    """
    try:
        queue.put("\n--- A registar no histórico de atividades ---")
        
        LOG_FILE_NAME = "controle_de_atividades.xlsx"
        SHEET_NAME = "Modelo"
        COLUMNS = [
            "Data", "Atividade Realizada", "Inscrição Municipal/CNPJ/CPF ou Nº do(s) Alvará(s) - CVCO", 
            "Verificações e Análises Realizadas", "Resultado", "Nº Processo ou Nº Certidão - CVCO", 
            "Nº DAM / IDD / AI / Denúncia", "Valor Original do ISS", "Valor Corrigido do ISS", "Horas Trabalhadas"
        ]

        # 1. Gather Data from Context
        data = context.get('data_por_extenso', pd.to_datetime('today').strftime("%d/%m/%Y"))
        
        # Ajusta o nome da atividade no log se for IDD
        idd_mode = context.get('idd_mode', False)
        atividade = "Operação IDD" if idd_mode else "Operação Receita"
        
        verificacoes = (
            "Análise da Sit cadastral, verificação das NFS-e, "
            "cruzamento das NFS-e com pagamentos realizados, "
            "Informação fiscal, Emissão de AUI´s de infração, "
            "termo de início e encerramento, envio via DEC, "
            "Ciência GTM e trâmite do protocolo"
        )
        resultado = "Protocolo concluído e AUI´s enviados para ciência"
        epaf = context.get('epaf_numero', '')
        
        # Get IDs from summary
        summary = context.get('summary', {})
        
        autos_list = summary.get('autos') or []
        auto_ids = [str(auto.get('numero', '')) for auto in autos_list]
        
        multa_info = summary.get('multa') or {}
        multa_num = multa_info.get('numero', None)
        
        all_ids = auto_ids[:]
        if multa_num:
            all_ids.append(str(multa_num))
        ids_str = ", ".join(filter(None, all_ids))

        # Get Values from Summary
        valor_original_total = sum(float(auto.get('iss_valor_original', 0.0)) for auto in autos_list)
        valor_corrigido_total = float(summary.get('total_geral_credito', 0.0))
        
        imu_cnpj = f"{context.get('imu', '')} / {context.get('cnpj', '')}" 
        horas = "" 

        # 2. Create DataFrame for the new row
        new_row_data = {
            "Data": [data],
            "Atividade Realizada": [atividade],
            "Inscrição Municipal/CNPJ/CPF ou Nº do(s) Alvará(s) - CVCO": [imu_cnpj], 
            "Verificações e Análises Realizadas": [verificacoes],
            "Resultado": [resultado],
            "Nº Processo ou Nº Certidão - CVCO": [epaf],
            "Nº DAM / IDD / AI / Denúncia": [ids_str],
            "Valor Original do ISS": [valor_original_total],
            "Valor Corrigido do ISS": [valor_corrigido_total],
            "Horas Trabalhadas": [horas]
        }
        new_row_df = pd.DataFrame(new_row_data, columns=COLUMNS)

        # 3. Write to Excel file
        if not os.path.exists(LOG_FILE_NAME):
            new_row_df.to_excel(LOG_FILE_NAME, index=False, sheet_name=SHEET_NAME)
            queue.put(f"✅ Histórico criado e atividade registada em {LOG_FILE_NAME}")
        else:
            try:
                # Carrega o workbook para verificar existência da folha
                workbook = openpyxl.load_workbook(LOG_FILE_NAME)
                if SHEET_NAME not in workbook.sheetnames:
                    # Se não existir, adiciona a folha usando ExcelWriter
                    with pd.ExcelWriter(LOG_FILE_NAME, engine='openpyxl', mode='a') as writer:
                        new_row_df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
                    queue.put(f"✅ Folha 'Modelo' criada e atividade registada.")
                else:
                    # Se existir, faz append
                    sheet = workbook[SHEET_NAME]
                    start_row = sheet.max_row
                    with pd.ExcelWriter(LOG_FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        new_row_df.to_excel(writer, index=False, header=False, sheet_name=SHEET_NAME, startrow=start_row)
                    queue.put(f"✅ Atividade registada com sucesso em {LOG_FILE_NAME}")
            
            except Exception as e:
                queue.put(f"❌ AVISO: Não foi possível registar a atividade (o ficheiro pode estar aberto): {e}")

    except Exception as e:
        queue.put(f"❌ AVISO: Falha crítica ao tentar registar atividade no Excel: {e}")
        logging.error(f"Failed to write to activity log: {traceback.format_exc()}")


def _generate_final_documents_task(queue, company_cnpj, master_filepath, final_data, preview_context, 
                                   temp_df_filepath, 
                                   numero_multa, dam_filepath, pgdas_folder_path, 
                                   base_output_dir_override, encerramento_version, company_imu,
                                   idd_mode=False):
    """
    Executa a lógica de geração de documentos dentro do subprocesso.
    """
    
    def emit(msg):
        queue.put(msg)

    logging.info("--- [SUBPROCESSO] Starting generate_final_documents ---")
    logging.info(f"[SUBPROCESSO] IDD Mode: {idd_mode}")

    try:
        emit("🔄 Carregando dados de faturas do ficheiro temporário...")
        # Carrega o DF do disco (passado via pickle pelo main thread)
        company_invoices_df = pd.read_pickle(temp_df_filepath) 
        logging.info(f"[SUBPROCESSO] DataFrame carregado com {len(company_invoices_df)} linhas.")
    except Exception as e:
        emit(f"❌ ERRO CRÍTICO: Falha ao carregar DataFrame do disco. Erro: {e}")
        raise e

    logging.info("[SUBPROCESSO] Building context...")
    
    # Passa idd_mode para o criador de contexto
    context = create_context_for_generation(
        master_filepath, company_cnpj,
        final_data, preview_context, 
        numero_multa, company_invoices_df,
        company_imu,
        idd_mode=idd_mode,
        dam_filepath=dam_filepath 
    )
    
    if not context:
        emit("❌ ERRO: Falha ao construir o contexto final.")
        return None

    try:
        if company_invoices_df is not None and not company_invoices_df.empty:
            if 'DATA EMISSÃO' in company_invoices_df.columns:
                # Ensure we have datetime objects
                dates = pd.to_datetime(company_invoices_df['DATA EMISSÃO'], errors='coerce').dropna()
                
                if not dates.empty:
                    min_year = int(dates.dt.year.min())
                    max_year = int(dates.dt.year.max())
                    
                    # Logic: 
                    # Same year: "janeiro a dezembro de 2022"
                    # Diff years: "janeiro de 2021 a dezembro de 2022"
                    if min_year == max_year:
                        full_period_str = f"janeiro a dezembro de {min_year}"
                    else:
                        full_period_str = f"janeiro de {min_year} a dezembro de {max_year}"
                    
                    # Force update all common variable names for period
                    context['periodo_fiscalizado'] = full_period_str
                    context['periodo'] = full_period_str 
                    context['periodo_extenso'] = full_period_str
                    
                    logging.info(f"📅 Período Fiscalizado ajustado para: {full_period_str}")
    except Exception as e:
        logging.error(f"Erro ao ajustar string de período fiscalizado: {e}")

    # --- CHANGED: DIRECTORY LOGIC ---
    file_prefix = "".join(filter(str.isdigit, company_cnpj))

    if base_output_dir_override:
        # If an override is provided (Batch mode), use it DIRECTLY.
        # Do NOT create a CNPJ subfolder.
        company_output_dir = base_output_dir_override
    else:
        # Default behavior: Use config dir and create CNPJ subfolder
        base_output_dir = get_output_dir()
        company_output_dir = os.path.join(base_output_dir, file_prefix)

    if not os.path.exists(company_output_dir):
        os.makedirs(company_output_dir)
    
    # Define output file paths
    output_inicio = os.path.join(company_output_dir, "termo_de_inicio.docx")
    output_relatorio = os.path.join(company_output_dir, "informacao_fiscal.docx")
    output_encerramento = os.path.join(company_output_dir, "termo_de_encerramento.docx")
    temp_file = os.path.join(company_output_dir, f"temp_{file_prefix}.docx")

    # --- Generate Main Word Documents ---
    emit("\n--- Gerando Documentos Principais ---")

    # ✅ --- FIX: Skip Start/End Terms if in IDD Mode ---
    if not idd_mode:
        logging.info("[SUBPROCESSO] Generating 'Termo de Início'...")
        inicio_template_path = os.path.normpath(get_template_inicio_path())
        generate_simple_document(context, inicio_template_path, output_inicio)
        emit(f"✅ Documento '{os.path.basename(output_inicio)}' gerado.")
        time.sleep(0.5) 

        template_path_encerramento = ""
        if encerramento_version == "AR":
            template_path_encerramento = os.path.normpath(get_template_encerramento_ar_path())
            logging.info("[SUBPROCESSO] Generating 'Termo de Encerramento (AR)'...")
        else:
            template_path_encerramento = os.path.normpath(get_template_encerramento_dec_path())
            logging.info("[SUBPROCESSO] Generating 'Termo de Encerramento (DEC)'...")

        generate_simple_document(context, template_path_encerramento, output_encerramento)
        emit(f"✅ Documento '{os.path.basename(output_encerramento)}' gerado.")
        time.sleep(0.5)
    else:
        emit("ℹ️ Modo IDD Ativo: Apenas a Informação Fiscal será gerada (Termos ignorados).")
    # ✅ --- END FIX ---

    logging.info("[SUBPROCESSO] Generating 'Relatório Final'...")
    relatorio_template_path = os.path.normpath(get_template_relatorio_path())
    generate_report(context, relatorio_template_path, output_relatorio, temp_file)
    emit(f"✅ Documento '{os.path.basename(output_relatorio)}' gerado.")
    time.sleep(0.5)
    
    # --- Generate Additional Detailed PDFs ---
    emit("\n--- Gerando Relatórios PDF Adicionais ---")
    logging.info("[SUBPROCESSO] Generating detailed PDFs...")

    generate_detailed_pdfs(
        context,
        company_invoices_df, 
        final_data,
        context, 
        company_output_dir,
        emit
    )
    logging.info("[SUBPROCESSO] Detailed PDFs generated.")

    emit(f"\n✨ Processo para {company_cnpj} concluído com sucesso!")

    # Limpa o ficheiro temporário do Word
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except Exception:
            pass
            
    return company_output_dir, context


def run_generation_task(queue, generation_args):
    """
    Função de entrada do subprocesso.
    """
    if multiprocessing.current_process().daemon:
        return

    shared_info = generation_args.get('shared_df_info', {})
    temp_df_filepath = shared_info.get('path', None)
    
    context_for_log = None
    output_directory = generation_args.get('output_dir', 'output')

    try:
        logging.info("[SUBPROCESSO] Tarefa de geração iniciada.")
        
        # Extrair todos os argumentos
        cnpj = generation_args['cnpj']
        master_path = generation_args['master_path']
        final_data = generation_args['final_data']
        preview_context = generation_args['preview_context']
        numero_multa = generation_args['numero_multa']
        dam_filepath = generation_args['dam_filepath']
        pgdas_folder_path = generation_args['pgdas_folder_path']
        output_dir = generation_args['output_dir']
        encerramento_version = generation_args['encerramento_version']
        company_imu = generation_args['company_imu']
        
        # Pega idd_mode com default False se não existir
        idd_mode = generation_args.get('idd_mode', False)

        # Executa a lógica principal
        result = _generate_final_documents_task(
            queue, cnpj, master_path, final_data, preview_context, 
            temp_df_filepath,
            numero_multa, dam_filepath, 
            pgdas_folder_path, output_dir, encerramento_version,
            company_imu,
            idd_mode=idd_mode 
        )
        
        if result is None:
            raise Exception("Falha ao gerar documentos (contexto vazio).")
            
        output_directory, context_for_log = result

        # Registra no Excel
        if context_for_log:
            _write_to_activity_log(queue, context_for_log)
        else:
            queue.put("❌ AVISO: Contexto final não retornado, log ignorado.")

        queue.put(("SUCCESS", output_directory))

    except BaseException as e:
        tb = traceback.format_exc()
        logging.critical(f"[SUBPROCESSO] CRASH: {e}\n{tb}")
        queue.put(f"❌ ERRO CRÍTICO NA GERAÇÃO: {e}\n{tb}")
        queue.put(("ERROR", tb))
        
    finally:
        # Limpeza do ficheiro pickle temporário
        if temp_df_filepath and os.path.exists(temp_df_filepath):
            try:
                os.remove(temp_df_filepath)
            except Exception:
                pass