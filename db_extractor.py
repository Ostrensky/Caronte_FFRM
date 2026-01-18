# --- app/ferramentas/db_extractor.py ---
# OTIMIZADO: Single SQL Query (Performance) -> Split Output (One file per year)

import pyodbc
import pandas as pd
import os
import re
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QThread

# --- SQL QUERY TEMPLATE FOR NOTAS (NFSE) ---
SQL_QUERY_NOTAS_TEMPLATE = """
SELECT
    -- 1. Header
    NFE.Num_Nota_Fiscal AS "NÚMERO",
    NFE.Num_Protocolo AS "CÓDIGO VERIFICAÇÃO",
    NFE.Idf_Lote_RPS AS "Nº LOTE",
    NFE.Num_RPS AS "Nº RPS",
    NFE.Dta_Emissao_Nota_Fiscal AS "DATA EMISSÃO",
    NFE.Dta_Conversao_RPS AS "DATA CONVERSÃO RPS",
    NFE.Dta_Canc_Nota_Fiscal AS "DT. CANCELAMENTO",

    -- 2. Payment Logic (Optimized)
    CASE 
        WHEN P.IsPaid = 1 AND I.HasInfraction = 1 THEN 'IDD'
        WHEN P.IsPaid = 1 THEN 'Sim'
        ELSE 'Não'
    END AS "PAGAMENTO",

    -- 3. Natureza da Operação
    CASE NFE.Ind_Natureza_Operacao_RPS
        WHEN 1 THEN 'TributacaoMunicipio'
        WHEN 2 THEN 'TributacaoForaMunicipio'
        WHEN 3 THEN 'Isencao'
        WHEN 4 THEN 'Imune'
        WHEN 5 THEN 'Exigibilidade suspensa por decisao judicial'
        WHEN 6 THEN 'Exigibilidade suspensa por procedimento [administrativo'
        ELSE CAST(NFE.Ind_Natureza_Operacao_RPS AS VARCHAR)
    END AS "NATUREZA DA OPERAÇÃO",

    -- 4. Values (Aggregated in V_AGG below)
    ISNULL(V_AGG.TotalValor, 0) AS "VALOR",
    ISNULL(V_AGG.TotalDesconto, 0) AS "DESCONTO INCONDICIONAL",
    ISNULL(V_AGG.TotalDeducao, 0) AS "VALOR DEDUÇÃO",
    ISNULL(V_AGG.MaxAliquota, 0) AS "ALÍQUOTA",

    -- 5. Provider
    NFE.Num_CNPJ_Prestador AS "CNPJ PRESTADOR",
    NFE.Nme_Prestador COLLATE DATABASE_DEFAULT AS "PRESTADOR",

    -- 6. Taker (Raw columns for Python processing)
    NFE.Num_CNPJ_Tomador AS "RAW_CNPJ_TOMADOR",
    NFE.Num_CPF_Tomador AS "RAW_CPF_TOMADOR",
    NFE.Nme_Tomador COLLATE DATABASE_DEFAULT AS "TOMADOR",

    -- 7. Regime (Updated Logic using Flg_Opt_Simples_Nac_RPS)
    CASE 
        WHEN NFE.Flg_Opt_Simples_Nac_RPS = 1 THEN 'Optante pelo Simples Nacional'
        ELSE 'Contribuinte sujeito a tributação normal'
    END AS "REGIME DE TRIBUTAÇÃO",

    -- 8. ISS Retained
    CASE 
        WHEN V_AGG.MaxRetido = 1 THEN 'Sim' 
        ELSE 'Não' 
    END AS "ISS RETIDO",

    -- 9. Description (XML Path kept for concatenation)
    STUFF((
        SELECT ' | ' + ISNULL(V.Des_Servico, '') COLLATE DATABASE_DEFAULT
        FROM [dbo].[ISSVNEValor_Nota_Eletronica] V
        WHERE V.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
        FOR XML PATH(''), TYPE).value('.', 'NVARCHAR(MAX)'), 1, 3, '') AS "DISCRIMINAÇÃO DOS SERVIÇOS",

    -- 10. Activity (XML Path kept for concatenation)
    STUFF((
        SELECT ', ' + CAST(V.Idf_Item_Lista_Servico AS VARCHAR(20)) COLLATE DATABASE_DEFAULT
        FROM [dbo].[ISSVNEValor_Nota_Eletronica] V
        WHERE V.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
        FOR XML PATH(''), TYPE).value('.', 'NVARCHAR(MAX)'), 1, 2, '') AS "CÓDIGO DA ATIVIDADE",

    -- 11. Address
    NFE.Nme_Cidade_ET COLLATE DATABASE_DEFAULT AS "CIDADE TOMADOR",
    (
        ISNULL(NFE.Nme_Logradouro_ET, '') COLLATE DATABASE_DEFAULT + ', ' + 
        ISNULL(NFE.Num_Endereco_ET, '') COLLATE DATABASE_DEFAULT + ' - ' + 
        ISNULL(NFE.Des_Complemento_ET, '') COLLATE DATABASE_DEFAULT + ' - ' + 
        ISNULL(NFE.Nme_Bairro_ET, '') COLLATE DATABASE_DEFAULT
    ) AS "ENDEREÇO TOMADOR"

FROM [dbo].[ISSNFENota_Fiscal_Eletronica] AS NFE

-- OPTIMIZATION 1: Single pass for values
OUTER APPLY (
    SELECT 
        SUM(V.Vlr_Servico) as TotalValor,
        SUM(V.Vlr_Desconto_Incondicionado) as TotalDesconto,
        SUM(V.Vlr_Deducao) as TotalDeducao,
        MAX(V.Vlr_Percentual_Aliquota) as MaxAliquota,
        MAX(CAST(V.Flg_ISS_Retido AS INT)) as MaxRetido
    FROM [dbo].[ISSVNEValor_Nota_Eletronica] V 
    WHERE V.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
) V_AGG

-- OPTIMIZATION 2: Single pass for Payment Status
OUTER APPLY (
    SELECT TOP 1 1 AS IsPaid
    FROM [dbo].[ISSIDAItens_Docto_Arrecadacao] IDA
    INNER JOIN [dbo].[ISSDARDocumento_Arrecadacao] DAR 
        ON IDA.Idf_Documento_Arrecadacao = DAR.Idf_Documento_Arrecadacao
    WHERE IDA.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
    AND DAR.Dta_Pagamento IS NOT NULL
) P

-- OPTIMIZATION 3: Single pass for Infraction Status
OUTER APPLY (
    SELECT TOP 1 1 AS HasInfraction
    FROM [dbo].[ISSNAFNotas_Auto_Infracao] NAF
    WHERE NAF.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
) I

WHERE
    NFE.Num_IM_Prestador COLLATE DATABASE_DEFAULT = ?
    AND YEAR(NFE.Dta_Emissao_Nota_Fiscal) IN ({year_placeholders})

ORDER BY
    NFE.Dta_Emissao_Nota_Fiscal,
    NFE.Num_Nota_Fiscal;
"""

# --- HELPER FUNCTIONS ---

def format_cnpj_str(val):
    if not val or val == '0' or pd.isna(val): return ""
    clean_val = re.sub(r'\D', '', str(val))
    clean_val = clean_val.zfill(14)
    return f"{clean_val[:2]}.{clean_val[2:5]}.{clean_val[5:8]}/{clean_val[8:12]}-{clean_val[12:]}"

def format_cpf_str(val):
    if not val or val == '0' or pd.isna(val): return ""
    clean_val = re.sub(r'\D', '', str(val))
    clean_val = clean_val.zfill(11)
    return f"{clean_val[:3]}.{clean_val[3:6]}.{clean_val[6:9]}-{clean_val[9:]}"

# --- DAMS-SPECIFIC FUNCTIONS (MULTI-YEAR QUERY) ---

def fetch_paid_dams(inscricao_municipal, years_list, conn, progress_callback):
    """
    Fetches DAMS for a LIST of years in a SINGLE query.
    Updated: Uses quoted string comparison to avoid 'varchar to int' conversion errors.
    """
    try:
        # Use str() to preserve value for string query, remove non-digits for safety if needed
        safe_im_public = re.sub(r'\D', '', str(inscricao_municipal)) 
        safe_years = [int(y) for y in years_list]
        if not safe_years: return None
    except ValueError:
        progress_callback.emit(f"  [DAMS] ❌ Erro: IM ou Ano inválidos.")
        return None

    try:
        cursor = conn.cursor()
        # --- STEP 1: RESOLVE PUBLIC IM -> INTERNAL ID ---
        sql_lookup_id = f"""
        SELECT TOP 1 DAR.Idf_Inscricao_Municipal
        FROM [dbo].[ISSDARDocumento_Arrecadacao] DAR
        INNER JOIN [dbo].[ISSIDAItens_Docto_Arrecadacao] IDA 
            ON DAR.Idf_Documento_Arrecadacao = IDA.Idf_Documento_Arrecadacao
        INNER JOIN [dbo].[ISSNFENota_Fiscal_Eletronica] NFE 
            ON IDA.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
        WHERE NFE.Num_IM_Prestador = '{safe_im_public}' 
        """
        
        cursor.execute(sql_lookup_id)
        row_id = cursor.fetchone()
        
        if not row_id:
            internal_id = safe_im_public
        else:
            internal_id = row_id[0]

        # --- STEP 2: MULTI-YEAR QUERY ---
        years_str = ",".join(map(str, safe_years))
        
        sql_query_dynamic = f"""
        SELECT DISTINCT
            DAR.Mes_Referencia_Pagamento,
            DAR.Ano_Referencia_Pagamento,
            CAST(DAR.Cod_Identificacao_Baixa AS VARCHAR(50)) AS "codigoVerificacao",
            DAR.Vlr_Total_Documento AS "totalRecolher",
            CAST(DAR.Mes_Referencia_Pagamento AS VARCHAR) + '/' + CAST(DAR.Ano_Referencia_Pagamento AS VARCHAR) AS "referenciaPagamento",
            DAR.Vlr_Receita AS "receita",
            ISNULL(DAR.Vlr_Deducao, 0) AS "desconto",
            CASE DAR.Idf_Reg_Tributo
                WHEN 197 THEN 'ISS Normal'
                WHEN 198 THEN 'ISS Retido na Fonte'
                ELSE 'Outros / ' + CAST(ISNULL(DAR.Idf_Reg_Tributo, 0) AS VARCHAR)
            END AS "tributo",
            STUFF((
                SELECT ', ' + CAST(NFE_SUB.Num_Nota_Fiscal AS VARCHAR)
                FROM [dbo].[ISSIDAItens_Docto_Arrecadacao] IDA_SUB
                INNER JOIN [dbo].[ISSNFENota_Fiscal_Eletronica] NFE_SUB 
                    ON IDA_SUB.Idf_Nota_Fiscal_Eletronica = NFE_SUB.Idf_Nota_Fiscal_Eletronica
                WHERE IDA_SUB.Idf_Documento_Arrecadacao = DAR.Idf_Documento_Arrecadacao
                ORDER BY NFE_SUB.Num_Nota_Fiscal
                FOR XML PATH(''), TYPE).value('.', 'NVARCHAR(MAX)'), 1, 2, '') AS "numerosDasNotas"
        FROM [dbo].[ISSDARDocumento_Arrecadacao] DAR
        WHERE 
            DAR.Idf_Inscricao_Municipal = '{internal_id}'
            AND DAR.Ano_Referencia_Pagamento IN ({years_str})
            AND DAR.Dta_Pagamento IS NOT NULL
            AND DAR.Vlr_Pagamento > 0
            AND DAR.Cod_Identificacao_Baixa IS NOT NULL
            AND CAST(DAR.Cod_Identificacao_Baixa AS VARCHAR) <> ''
            AND CAST(DAR.Cod_Identificacao_Baixa AS VARCHAR) <> '-'
        ORDER BY DAR.Ano_Referencia_Pagamento, DAR.Mes_Referencia_Pagamento;
        """

        df = pd.read_sql(sql_query_dynamic, conn)

        if df.empty:
            return None
        
        # --- FAST DATA CLEANING ---
        df['codigoVerificacao'] = df['codigoVerificacao'].astype(str).str.strip()
        df = df[~df['codigoVerificacao'].isin(['', '-', 'None', 'none'])]
        
        float_cols = ['totalRecolher', 'receita', 'desconto']
        for col in float_cols:
             df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        
        df['numerosDasNotas'] = df['numerosDasNotas'].fillna("")

        progress_callback.emit(f"    [DAMS] {len(df)} registros encontrados (Intervalo Completo).")
        return df

    except Exception as e:
        progress_callback.emit(f"    [DAMS] ❌ Erro BD: {e}")
        return None

def export_to_csv(df, filename, progress_callback):
    if df is None or df.empty:
        # progress_callback.emit(f"  [DAMS] Nenhum dado para salvar.")
        return

    progress_callback.emit(f"  [DAMS] Salvando: {os.path.basename(filename)}...")
    
    cols_order = [
        "codigoVerificacao", "totalRecolher", "referenciaPagamento", 
        "receita", "desconto", "tributo", "numerosDasNotas"
    ]
    
    for col in cols_order:
        if col not in df.columns:
            df[col] = 0.0 if col in ["totalRecolher", "receita", "desconto"] else ""

    df = df[cols_order]

    try:
        df.to_csv(filename, index=False, sep=',', decimal='.', encoding='utf-8-sig')
    except Exception as e:
        progress_callback.emit(f"  [DAMS] ❌ Erro ao salvar CSV: {e}")

# --- NOTAS (NFSE)-SPECIFIC FUNCTIONS (MULTI-YEAR QUERY) ---

def fetch_invoice_data(inscricao_municipal, years_list, conn, progress_callback):
    """
    Fetches NFSE for a LIST of years in a SINGLE query using 'IN' clause.
    """
    try:
        placeholders = ', '.join(['?'] * len(years_list))
        query = SQL_QUERY_NOTAS_TEMPLATE.format(year_placeholders=placeholders)
        
        # Ensure IM is string to match VARCHAR column type
        params = [str(inscricao_municipal)] + years_list
        
        df = pd.read_sql(query, conn, params=params)

        if df.empty:
            return None

        # --- FAST DATA CLEANING ---
        # Keep datetime object for splitting, convert later if needed for display
        df['DATA EMISSÃO RAW'] = pd.to_datetime(df['DATA EMISSÃO'], errors='coerce')
        
        # Format columns for output
        date_cols = ["DATA EMISSÃO", "DT. CANCELAMENTO"]
        for col in date_cols:
            # We update the string column but keep RAW for filtering
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%m/%d/%y %H:%M').fillna(pd.NaT)

        float_cols = ["VALOR", "DESCONTO INCONDICIONAL", "VALOR DEDUÇÃO", "ALÍQUOTA"]
        for col in float_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        # Logic for CNPJ/CPF TOMADOR
        def determine_taker_doc(row):
            cnpj = str(row['RAW_CNPJ_TOMADOR']).strip()
            cpf = str(row['RAW_CPF_TOMADOR']).strip()
            if cnpj and cnpj not in ['0', '', 'None']:
                return format_cnpj_str(cnpj)
            elif cpf and cpf not in ['0', '', 'None']:
                return format_cpf_str(cpf)
            return ""

        df['CNPJ/CPF TOMADOR'] = df.apply(determine_taker_doc, axis=1)
        df['CNPJ PRESTADOR'] = df['CNPJ PRESTADOR'].apply(format_cnpj_str)

        progress_callback.emit(f"    [NFSE] {len(df)} notas processadas (Intervalo Completo).")
        return df

    except Exception as e:
        progress_callback.emit(f"    [NFSE] ❌ Erro BD: {e}")
        return None

def export_to_excel(df, filename, progress_callback):
    if df is None or df.empty:
        # progress_callback.emit(f"  [NFSE] Nenhum dado para salvar.")
        return

    progress_callback.emit(f"  [NFSE] Salvando: {os.path.basename(filename)}...")
    
    # Drop temp column if exists
    if 'DATA EMISSÃO RAW' in df.columns:
        df = df.drop(columns=['DATA EMISSÃO RAW'])

    cols_order = [
        "NÚMERO", "CÓDIGO VERIFICAÇÃO", "Nº LOTE", "Nº RPS", 
        "DATA EMISSÃO", "DATA CONVERSÃO RPS", "DT. CANCELAMENTO", 
        "VALOR", "DESCONTO INCONDICIONAL", "VALOR DEDUÇÃO", "ALÍQUOTA", 
        "CNPJ PRESTADOR", "PRESTADOR", "CNPJ/CPF TOMADOR", "TOMADOR", 
        "REGIME DE TRIBUTAÇÃO", "NATUREZA DA OPERAÇÃO", "ISS RETIDO", 
        "DISCRIMINAÇÃO DOS SERVIÇOS", "CÓDIGO DA ATIVIDADE", 
        "PAGAMENTO", "CIDADE TOMADOR", "ENDEREÇO TOMADOR"
    ]
    
    for col in cols_order:
        if col not in df.columns:
            df[col] = 0.0 if col in ["VALOR", "DESCONTO INCONDICIONAL", "VALOR DEDUÇÃO", "ALÍQUOTA"] else ""
            
    df = df[cols_order]

    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=2, sheet_name='Relatorio_NFSE')
            
            worksheet = writer.sheets['Relatorio_NFSE']
            
            numeric_columns = ["VALOR", "DESCONTO INCONDICIONAL", "VALOR DEDUÇÃO", "ALÍQUOTA"]
            col_map = {col: idx+1 for idx, col in enumerate(cols_order)}
            
            for col_name in numeric_columns:
                col_idx = col_map.get(col_name)
                if col_idx:
                    col_letter = get_column_letter(col_idx)
                    for cell in worksheet[col_letter]:
                        if cell.row >= 4: 
                            cell.number_format = '#,##0.00'

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
    except Exception as e:
        progress_callback.emit(f"  [NFSE] ❌ Erro ao salvar Excel: {e}")

# --- MAIN WORKER FUNCTION (SPLIT OUTPUT) ---

def fetch_paid_das(inscricao_municipal, years_list, conn, progress_callback):
    """
    Fetches DAS (Simples Nacional) for a LIST of years.
    Based on dbo.ISSDASDoc_Arrecadacao_Simples.
    """
    try:
        # Data cleaning for inputs
        safe_im_public = re.sub(r'\D', '', str(inscricao_municipal)) 
        safe_years = [int(y) for y in years_list]
        if not safe_years: return None
    except ValueError:
        progress_callback.emit(f"  [DAS] ❌ Erro: IM ou Ano inválidos.")
        return None

    try:
        cursor = conn.cursor()
        
        # --- STEP 1: RESOLVE PUBLIC IM -> INTERNAL ID ---
        # Reuse logic from DAMS/NFSE
        sql_lookup_id = f"""
        SELECT TOP 1 DAR.Idf_Inscricao_Municipal
        FROM [dbo].[ISSDARDocumento_Arrecadacao] DAR
        INNER JOIN [dbo].[ISSIDAItens_Docto_Arrecadacao] IDA 
            ON DAR.Idf_Documento_Arrecadacao = IDA.Idf_Documento_Arrecadacao
        INNER JOIN [dbo].[ISSNFENota_Fiscal_Eletronica] NFE 
            ON IDA.Idf_Nota_Fiscal_Eletronica = NFE.Idf_Nota_Fiscal_Eletronica
        WHERE NFE.Num_IM_Prestador = '{safe_im_public}' 
        """
        
        internal_id = safe_im_public
        try:
            cursor.execute(sql_lookup_id)
            row_id = cursor.fetchone()
            if row_id:
                internal_id = row_id[0]
        except Exception:
            pass # Fallback to safe_im_public

        # --- STEP 2: QUERY THE DAS TABLE ---
        years_str = ",".join(map(str, safe_years))
        
        sql_query_das = f"""
        SELECT
            DAS.Mes_Referencia_Pagamento,
            DAS.Ano_Referencia_Pagamento,
            CAST(DAS.Cod_Identificacao_Baixa AS VARCHAR(50)) AS "codigoVerificacao",
            DAS.Vlr_Pagamento AS "valorPago",
            ISNULL(DAS.Vlr_Multa, 0) AS "multa",
            ISNULL(DAS.Vlr_Juros, 0) AS "juros",
            CAST(DAS.Mes_Referencia_Pagamento AS VARCHAR) + '/' + CAST(DAS.Ano_Referencia_Pagamento AS VARCHAR) AS "referenciaPagamento",
            DAS.Dta_Pagamento AS "dataPagamento"
        FROM [dbo].[ISSDASDoc_Arrecadacao_Simples] DAS
        WHERE 
            DAS.Idf_Inscricao_Municipal = '{internal_id}'
            AND DAS.Ano_Referencia_Pagamento IN ({years_str})
            AND DAS.Dta_Pagamento IS NOT NULL
            AND DAS.Vlr_Pagamento > 0
            AND DAS.Cod_Identificacao_Baixa IS NOT NULL
            AND CAST(DAS.Cod_Identificacao_Baixa AS VARCHAR) <> ''
            AND CAST(DAS.Cod_Identificacao_Baixa AS VARCHAR) <> '-'
        ORDER BY DAS.Ano_Referencia_Pagamento, DAS.Mes_Referencia_Pagamento;
        """

        df = pd.read_sql(sql_query_das, conn)

        if df.empty:
            return None
        
        # --- FAST DATA CLEANING ---
        df['codigoVerificacao'] = df['codigoVerificacao'].astype(str).str.strip()
        float_cols = ['valorPago', 'multa', 'juros']
        for col in float_cols:
             df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        df['dataPagamento'] = pd.to_datetime(df['dataPagamento'], errors='coerce').dt.strftime('%d/%m/%Y')

        progress_callback.emit(f"    [DAS] {len(df)} registros encontrados.")
        return df

    except Exception as e:
        progress_callback.emit(f"    [DAS] ❌ Erro BD: {e}")
        return None

def export_das_to_csv(df, filename, progress_callback):
    if df is None or df.empty: return

    progress_callback.emit(f"  [DAS] Salvando: {os.path.basename(filename)}...")
    
    cols_order = ["codigoVerificacao", "referenciaPagamento", "dataPagamento", "valorPago", "multa", "juros"]
    
    for col in cols_order:
        if col not in df.columns:
            df[col] = 0.0 if col in ["valorPago", "multa", "juros"] else ""

    df = df[cols_order]

    try:
        df.to_csv(filename, index=False, sep=',', decimal='.', encoding='utf-8-sig')
    except Exception as e:
        progress_callback.emit(f"  [DAS] ❌ Erro ao salvar CSV: {e}")

# --- UPDATED MAIN WORKER FUNCTION ---

# ⚠️ UPDATE SIGNATURE TO ACCEPT do_das
def run_db_extraction(target_list, years_list, do_dams, do_nfse, do_das, progress_callback):
    
    def check_stop_flag():
        current_thread = QThread.currentThread()
        if hasattr(current_thread, 'check_stop') and current_thread.check_stop():
            return True
        return False
    
    if not target_list:
        progress_callback.emit("❌ Lista de tarefas vazia.")
        return

    if not isinstance(years_list, list):
        years_list = [years_list]
    
    years_list = sorted([int(y) for y in years_list])
    
    # Settings (Credentials kept same as original)
    server = '172.19.210.187,1433'
    database = 'ISS_CURITIBA_RELATORIOS'
    username = 'vostrensky'
    password = 'T$&KzpUzUQ@yH4jchh' 

    CONN_STRING = (
        f'DRIVER={{SQL Server}};' 
        f'SERVER={server};'
        f'DATABASE={database};'
        f'UID={username};'
        f'PWD={password};'
        f'TrustServerCertificate=yes;'
    )

    total_dirs = len(target_list)
    str_years = ", ".join(map(str, years_list))
    
    progress_callback.emit(f"--- Iniciando Extrator BD (Multi-Ano) ---")
    progress_callback.emit(f"Alvos: {total_dirs} | Período: {str_years}")
    progress_callback.emit(f"Opções: DAMS={'Sim' if do_dams else 'Não'} | NFSE={'Sim' if do_nfse else 'Não'} | DAS={'Sim' if do_das else 'Não'}")

    try:
        progress_callback.emit("🔌 Conectando ao Banco de Dados...")
        with pyodbc.connect(CONN_STRING) as conn:
            
            for i, item in enumerate(target_list):
                
                if check_stop_flag():
                    progress_callback.emit("🛑 Parada solicitada. Encerrando extração.")
                    break 
                
                subdir_path = item['path']
                entry_name = item['name']
                imu = item['imu']
                
                progress_callback.emit(f"\n--- Processando {i+1}/{total_dirs}: {entry_name} ({imu}) ---")
                
                # 1. PROCESS DAMS
                if do_dams:
                    try:
                        full_dams_df = fetch_paid_dams(imu, years_list, conn, progress_callback)
                        if full_dams_df is not None and not full_dams_df.empty:
                            unique_years_in_data = full_dams_df['Ano_Referencia_Pagamento'].unique()
                            for year in sorted(unique_years_in_data):
                                if check_stop_flag(): break
                                year_df = full_dams_df[full_dams_df['Ano_Referencia_Pagamento'] == year].copy()
                                fname = f"Relatorio_DAMS_{imu}_{year}.csv"
                                export_to_csv(year_df, os.path.join(subdir_path, fname), progress_callback)
                        else:
                            progress_callback.emit("    [DAMS] Nada encontrado.")
                    except Exception as e:
                        progress_callback.emit(f"    ❌ Erro DAMS: {e}")
                
                # 2. PROCESS NFSE
                if do_nfse:
                    try:
                        full_nfse_df = fetch_invoice_data(imu, years_list, conn, progress_callback)
                        if full_nfse_df is not None and not full_nfse_df.empty:
                            full_nfse_df['year_temp'] = full_nfse_df['DATA EMISSÃO RAW'].dt.year
                            unique_years_in_data = full_nfse_df['year_temp'].dropna().unique()
                            for year in sorted(unique_years_in_data):
                                if check_stop_flag(): break
                                year = int(year)
                                year_df = full_nfse_df[full_nfse_df['year_temp'] == year].copy()
                                year_df = year_df.drop(columns=['year_temp'])
                                fname = f"Relatorio_NFSE_{imu}_{year}.xlsx"
                                export_to_excel(year_df, os.path.join(subdir_path, fname), progress_callback)
                        else:
                             progress_callback.emit("    [NFSE] Nada encontrado.")
                    except Exception as e:
                        progress_callback.emit(f"    ❌ Erro NFSE: {e}")

                # 3. PROCESS DAS (SIMPLES) - NEW BLOCK
                if do_das:
                    try:
                        full_das_df = fetch_paid_das(imu, years_list, conn, progress_callback)
                        if full_das_df is not None and not full_das_df.empty:
                            unique_years_in_data = full_das_df['Ano_Referencia_Pagamento'].unique()
                            for year in sorted(unique_years_in_data):
                                if check_stop_flag(): break
                                year_df = full_das_df[full_das_df['Ano_Referencia_Pagamento'] == year].copy()
                                fname = f"Relatorio_DAS_{imu}_{year}.csv"
                                export_das_to_csv(year_df, os.path.join(subdir_path, fname), progress_callback)
                        else:
                            progress_callback.emit("    [DAS] Nada encontrado.")
                    except Exception as e:
                        progress_callback.emit(f"    ❌ Erro DAS: {e}")

    except pyodbc.Error as e:
        progress_callback.emit(f"❌ Erro Crítico de Conexão: {e}")
    except Exception as e:
        progress_callback.emit(f"❌ Erro Inesperado: {e}")
    
    if not check_stop_flag():
        progress_callback.emit("\n--- --- --- --- --- ---")
        progress_callback.emit("✅ Processo Finalizado.")
    else:
        progress_callback.emit("\n🛑 Interrompido pelo usuário.")